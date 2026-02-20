#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import io
import re
import sys
import time
import argparse
import warnings
from datetime import datetime
from pathlib import Path

# windows terminals default to cp1252 which breaks unicode prints
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from deep_translator import GoogleTranslator

warnings.filterwarnings("ignore")

DEFAULT_INPUT  = Path("Outscraper-20241216130344s54.xlsx")
DEFAULT_OUTPUT = Path(f"experience_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

# 7 target languages (English is the base, so 8 total)
TARGET_LANGUAGES = {
    "Spanish":    "es",
    "French":     "fr",
    "German":     "de",
    "Japanese":   "ja",
    "Chinese":    "zh-CN",
    "Arabic":     "ar",
    "Portuguese": "pt",
}

# if any of these appear in the 5 words before a keyword hit, flip its sentiment
NEGATION_WORDS = frozenset({
    "not", "no", "never", "neither", "nor", "without", "lacks", "lack",
    "don't", "doesnt", "doesn't", "didnt", "didn't",
    "won't", "wont", "wouldn't", "wouldnt",
    "can't", "cant", "cannot",
    "wasn't", "wasnt", "weren't", "werent",
    "isn't", "isnt", "aren't", "arent",
    "haven't", "havent", "hasn't", "hasnt", "hadn't", "hadnt",
    "hardly", "barely", "scarcely",
})

# keyword patterns grouped by experience category
# * = wildcard (matches word endings), spaces match hyphens too
# negation handling is automatic, so no need to add "not X" as a separate positive
EXPERIENCE_KEYWORDS = {
    "Recommendation": {
        "positive": [
            "recommend*",
            "must visit",
            "must-visit",
            "must see",
            "must-see",
            "worth visit*",
            "worth a visit",
            "worth the trip",
            "worth going",
            "must go",
            "go here",
            "highly suggest*",
            "check it out",
        ],
        "negative": [
            "avoid*",
            "stay away",
            "skip this",
            "skip it",
            "waste of time",
            "waste your time",
            "not worth visit*",
            "don't bother",
            "do not bother",
        ],
    },

    "Overall Impression": {
        "positive": [
            "amaz*",
            "wonderful*",
            "fantastic*",
            "excellent*",
            "incredib*",
            "outstanding*",
            "spectacular*",
            "breathtaking*",
            "stunning*",
            "magnificent*",
            "brilliant*",
            "superb*",
            "awesome*",
            "loved it",
            "love it",
            "love this",
            "loved this",
            "perfect*",
            "beautiful*",
            "gorgeous*",
            "marvelous*",
            "fabulous*",
            "phenomenal*",
        ],
        "negative": [
            "terrible*",
            "awful*",
            "horrible*",
            "dreadful*",
            "disappoint*",
            "disgrace*",
            "appalling*",
            "pathetic*",
            "atrocious*",
            "dismal*",
            "mediocre*",
            "underwhelm*",
            "let down",
            "letdown*",
        ],
    },

    "Value for Money": {
        "positive": [
            "worth it",
            "worth the money",
            "worth the price",
            "worth every penny",
            "good value",
            "great value",
            "excellent value",
            "affordable*",
            "budget friendly",
            "fair price*",
            "well priced",
            "value for money",
            "reasonable price*",
            "reasonably priced",
            "cheap*",
            "inexpensive*",
        ],
        "negative": [
            "overpriced*",
            "too expensive",
            "very expensive",
            "pricey*",
            "costly*",
            "rip off",
            "ripoff*",
            "ripped off",
            "poor value",
            "waste of money",
            "not worth the money",
            "not worth the price",
        ],
    },

    "Service Quality": {
        "positive": [
            "friendly staff*",
            "helpful staff*",
            "great service*",
            "excellent service*",
            "good service*",
            "amazing service*",
            "professional*",
            "attentive*",
            "courteous*",
            "welcoming*",
            "polite*",
            "knowledgeable*",
            "accommodating*",
            "staff were great",
            "staff was great",
            "staff were friendly",
            "staff was friendly",
        ],
        "negative": [
            "rude*",
            "unhelpful*",
            "unfriendly*",
            "poor service*",
            "bad service*",
            "terrible service*",
            "horrible service*",
            "slow service*",
            "bad staff*",
            "rude staff*",
            "staff were rude",
            "staff was rude",
            "dismissive*",
            "ignorant*",
        ],
    },

    "Atmosphere": {
        "positive": [
            "great atmosphere*",
            "amazing atmosphere*",
            "wonderful atmosphere*",
            "beautiful atmosphere*",
            "lovely atmosphere*",
            "great ambiance*",
            "great ambience*",
            "peaceful*",
            "relaxing*",
            "cozy*",
            "cosy*",
            "charming*",
            "romantic*",
            "vibrant*",
            "magical*",
            "serene*",
            "tranquil*",
            "enchanting*",
            "lively*",
        ],
        "negative": [
            "bad atmosphere*",
            "poor atmosphere*",
            "terrible atmosphere*",
            "no atmosphere*",
            "noisy*",
            "too loud*",
            "chaotic*",
            "depressing*",
            "lifeless*",
            "cramped*",
            "stuffy*",
            "claustrophobic*",
        ],
    },

    "Cleanliness": {
        "positive": [
            "very clean*",
            "super clean*",
            "spotless*",
            "well maintained*",
            "well-maintained*",
            "immaculate*",
            "pristine*",
            "hygienic*",
            "tidy*",
        ],
        "negative": [
            "dirty*",
            "filthy*",
            "messy*",
            "disgusting*",
            "unhygienic*",
            "grimy*",
            "smelly*",
            "stinky*",
            "run down*",
            "run-down*",
            "neglected*",
            "poorly maintained*",
        ],
    },

    "Crowd & Wait Times": {
        "positive": [
            "short wait*",
            "no wait*",
            "no queue*",
            "no line*",
            "quick entry*",
            "fast entry*",
            "easy entry*",
            "not too crowded*",
            "quiet*",
        ],
        "negative": [
            "crowded*",
            "overcrowded*",
            "packed*",
            "long wait*",
            "long queue*",
            "long line*",
            "very busy*",
            "too busy*",
            "no space*",
        ],
    },

    "Uniqueness & Memorability": {
        "positive": [
            "unique*",
            "one of a kind",
            "one-of-a-kind",
            "special*",
            "iconic*",
            "historic*",
            "cultural*",
            "unforgettable*",
            "memorable*",
            "once in a lifetime",
            "bucket list*",
            "hidden gem*",
            "gem of",
        ],
        "negative": [
            "overrated*",
            "nothing special",
            "boring*",
            "generic*",
            "touristy*",
            "gimmick*",
            "ordinary*",
            "not unique*",
        ],
    },
}


def _pattern_to_regex(pattern: str) -> re.Pattern:
    # split on * so we can escape each literal chunk separately,
    # then rejoin with \w* to handle word endings
    segments = pattern.split("*")
    escaped_segments = []
    for seg in segments:
        e = re.escape(seg)
        e = e.replace(r"\ ", r"[\s\-]+")
        e = e.replace(r"\-", r"[\s\-]+")
        escaped_segments.append(e)
    regex_body = r"\w*".join(escaped_segments)
    # lookarounds work better than \b when the pattern itself ends in \w*
    return re.compile(r"(?<!\w)" + regex_body + r"(?!\w)", re.IGNORECASE)


# compile everything upfront so we're not rebuilding regexes on every row
_COMPILED: dict[str, dict[str, list[tuple[str, re.Pattern]]]] = {}
for _cat, _sents in EXPERIENCE_KEYWORDS.items():
    _COMPILED[_cat] = {}
    for _sent, _pats in _sents.items():
        _COMPILED[_cat][_sent] = [(_p, _pattern_to_regex(_p)) for _p in _pats]


def _has_negation(text: str, match_start: int, window: int = 60) -> bool:
    # grab the 60 chars before the match and check the last 5 words
    snippet = text[max(0, match_start - window): match_start]
    words = re.findall(r"\b\w+(?:'\w+)?\b", snippet.lower())
    return bool(set(words[-5:]) & NEGATION_WORDS)


def analyze_text(text: str) -> dict[str, dict[str, int]]:
    if not isinstance(text, str) or not text.strip():
        return {}

    results: dict[str, dict[str, int]] = {}
    # track spans per category so the same phrase doesn't get counted twice
    cat_spans: dict[str, set[tuple[int, int]]] = {}

    for cat, sents in _COMPILED.items():
        pos = neg = 0
        used = cat_spans.setdefault(cat, set())
        for sentiment, patterns in sents.items():
            for _raw, regex in patterns:
                for m in regex.finditer(text):
                    span = (m.start(), m.end())
                    if span in used:
                        continue
                    used.add(span)
                    negated = _has_negation(text, m.start())
                    effective = sentiment if not negated else (
                        "negative" if sentiment == "positive" else "positive"
                    )
                    if effective == "positive":
                        pos += 1
                    else:
                        neg += 1
        if pos > 0 or neg > 0:
            results[cat] = {"positive": pos, "negative": neg}

    return results


def experience_score(pos: int, neg: int) -> float | None:
    # returns None when there are no keyword hits at all
    if pos + neg == 0:
        return None
    return round(50.0 + (pos - neg) / (pos + neg) * 50.0, 1)


def _clean_for_translation(pattern: str) -> str:
    return re.sub(r"\*+", "", pattern).strip()


def translate_keywords(
    patterns: list[str],
    target_langs: dict[str, str],
    pause: float = 0.35,
) -> dict[str, dict[str, str]]:
    # dedupe first so we don't translate the same phrase multiple times
    # (lots of patterns share the same cleaned string after removing *)
    phrase_to_patterns: dict[str, list[str]] = {}
    for pat in patterns:
        clean = _clean_for_translation(pat)
        phrase_to_patterns.setdefault(clean, []).append(pat)

    unique_phrases = list(phrase_to_patterns.keys())
    total_calls = len(unique_phrases) * len(target_langs)
    print(f"  Translating {len(unique_phrases)} unique phrases x "
          f"{len(target_langs)} languages = {total_calls} API calls ...")
    print(f"  (estimated ~{total_calls * pause / 60:.1f} min)\n")

    phrase_translations: dict[str, dict[str, str]] = {}
    done = 0
    for phrase in unique_phrases:
        phrase_translations[phrase] = {"English": phrase}
        for lang_name, lang_code in target_langs.items():
            try:
                translated = GoogleTranslator(
                    source="auto", target=lang_code
                ).translate(phrase)
                phrase_translations[phrase][lang_name] = translated or phrase
            except Exception as exc:
                phrase_translations[phrase][lang_name] = f"[error: {exc}]"
            done += 1
            time.sleep(pause)  # google translate doesn't like being hammered

        print(f"  [{done:>4}/{total_calls}] {phrase}")

    # map results back to the original patterns (with * intact)
    result: dict[str, dict[str, str]] = {}
    for pat in patterns:
        clean = _clean_for_translation(pat)
        result[pat] = phrase_translations.get(clean, {"English": clean})

    return result


# Excel colour/style constants
_H_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark navy
_SH_FILL = PatternFill("solid", fgColor="2E75B6")   # medium blue
_ALT     = PatternFill("solid", fgColor="D6E4F0")   # light blue alternating rows
_WHITE   = PatternFill("solid", fgColor="FFFFFF")
_GREEN   = PatternFill("solid", fgColor="C6EFCE")
_RED     = PatternFill("solid", fgColor="FFC7CE")
_YELLOW  = PatternFill("solid", fgColor="FFEB9C")

_H_FONT  = Font(bold=True, color="FFFFFF", size=11)
_B_FONT  = Font(size=10)
_THIN    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill      = _H_FILL
        cell.font      = _H_FONT
        cell.alignment = _CENTER
        cell.border    = _THIN


def _style_data_row(ws, row: int, ncols: int, alt: bool) -> None:
    fill = _ALT if alt else _WHITE
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill      = fill
        cell.font      = _B_FONT
        cell.border    = _THIN
        cell.alignment = _CENTER


def _auto_col_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col),
            default=0
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 52)


def build_report(
    summary_df:  pd.DataFrame,
    category_df: pd.DataFrame,
    translations: dict[str, dict[str, str]],
    output_path: Path,
) -> None:
    cats      = list(EXPERIENCE_KEYWORDS.keys())
    lang_cols = ["English"] + list(TARGET_LANGUAGES.keys())

    wb = openpyxl.Workbook()

    # sheet 1 - one row per attraction with overall score + per-category counts
    ws1 = wb.active
    ws1.title = "Attraction Summary"

    hdr1 = (
        ["Attraction Name", "Total Reviews", "Reviews w/ Text", "Avg Rating",
         "Overall Score (0-100)"]
        + [f"{c} (+)" for c in cats]
        + [f"{c} (-)" for c in cats]
    )
    ws1.append(hdr1)
    _style_header_row(ws1, 1, len(hdr1))
    ws1.row_dimensions[1].height = 55
    ws1.freeze_panes = "B2"

    for i, (_, r) in enumerate(summary_df.iterrows(), start=2):
        score = r.get("experience_score")
        row_data = [
            r["name"],
            r["total_reviews"],
            r["reviews_with_text"],
            r.get("avg_rating"),
            score,
        ] + [r.get(f"{c}_pos", 0) for c in cats] + [r.get(f"{c}_neg", 0) for c in cats]
        ws1.append(row_data)
        _style_data_row(ws1, i, len(hdr1), alt=(i % 2 == 0))
        ws1.cell(i, 1).alignment = _LEFT
        sc = ws1.cell(i, 5)
        if score is None:
            pass
        elif score >= 70:
            sc.fill = _GREEN
        elif score <= 40:
            sc.fill = _RED
        else:
            sc.fill = _YELLOW

    _auto_col_width(ws1)

    # sheet 2 - broken down by attraction x category
    ws2 = wb.create_sheet("Category Breakdown")
    hdr2 = ["Attraction", "Category", "Positive Hits", "Negative Hits",
            "Category Score (0-100)", "Sentiment"]
    ws2.append(hdr2)
    _style_header_row(ws2, 1, len(hdr2))
    ws2.row_dimensions[1].height = 40
    ws2.freeze_panes = "C2"

    for i, (_, r) in enumerate(
        category_df.sort_values(["name", "category"]).iterrows(), start=2
    ):
        score = experience_score(int(r["positive"]), int(r["negative"]))
        if score is None:
            sentiment = "-"
        elif score >= 60:
            sentiment = "Positive"
        elif score <= 40:
            sentiment = "Negative"
        else:
            sentiment = "Neutral"

        ws2.append([r["name"], r["category"],
                    int(r["positive"]), int(r["negative"]),
                    score, sentiment])
        _style_data_row(ws2, i, len(hdr2), alt=(i % 2 == 0))
        ws2.cell(i, 1).alignment = _LEFT
        ws2.cell(i, 2).alignment = _LEFT
        sc = ws2.cell(i, 5)
        if score is None:
            pass
        elif score >= 60:
            sc.fill = _GREEN
        elif score <= 40:
            sc.fill = _RED
        else:
            sc.fill = _YELLOW

    _auto_col_width(ws2)

    # sheet 3 - keyword translation lookup table
    ws3 = wb.create_sheet("Keyword Translations")
    hdr3 = ["Category", "Sentiment", "Pattern"] + lang_cols
    ws3.append(hdr3)
    _style_header_row(ws3, 1, len(hdr3))
    ws3.row_dimensions[1].height = 40
    ws3.freeze_panes = "D2"

    row_idx = 2
    for cat, sents in EXPERIENCE_KEYWORDS.items():
        for sentiment, patterns in sents.items():
            for pat in patterns:
                trans = translations.get(pat, {"English": _clean_for_translation(pat)})
                row_data = (
                    [cat, sentiment.capitalize(), pat]
                    + [trans.get(lang, "") for lang in lang_cols]
                )
                ws3.append(row_data)
                _style_data_row(ws3, row_idx, len(hdr3), alt=(row_idx % 2 == 0))
                for c in range(1, len(hdr3) + 1):
                    ws3.cell(row_idx, c).alignment = _LEFT
                row_idx += 1

    _auto_col_width(ws3)

    wb.save(output_path)
    print(f"\n  [OK] Report saved -> {output_path.resolve()}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Outscraper experience analysis with NLP, scoring, and Excel export."
    )
    parser.add_argument(
        "--input", default=str(DEFAULT_INPUT),
        help="Path to the Outscraper Excel (.xlsx) file"
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUTPUT),
        help="Path for the output Excel report"
    )
    parser.add_argument(
        "--no-translate", action="store_true",
        help="Skip Google Translate step (faster; keywords shown in English only)"
    )
    args = parser.parse_args()

    input_file  = Path(args.input)
    output_file = Path(args.output)

    print("=" * 70)
    print("  Outscraper Experience Analysis")
    print("=" * 70)

    # load data
    print(f"\n[1/5] Loading {input_file} ...")
    if not input_file.exists():
        sys.exit(f"  ERROR: file not found -> {input_file}")

    # the actual data is on the second sheet - sheet 0 is just a stub
    import openpyxl as _oxl
    _wb = _oxl.load_workbook(input_file, read_only=True)
    _sheets = _wb.sheetnames
    _wb.close()
    print(f"      Sheets found: {_sheets}")

    sheet_idx = 0
    for _i, _sname in enumerate(_sheets):
        _test = pd.read_excel(input_file, engine="openpyxl", sheet_name=_i, nrows=2)
        if "name" in _test.columns:
            sheet_idx = _i
            break
    print(f"      Using sheet index {sheet_idx}: '{_sheets[sheet_idx]}'")

    df = pd.read_excel(input_file, engine="openpyxl", sheet_name=sheet_idx)
    print(f"      {len(df):,} rows x {len(df.columns)} columns loaded.")

    df = df.dropna(subset=["name"])
    print(f"      {len(df):,} rows after dropping empty names.")

    TEXT_COL   = "review_text"
    RATING_COL = "review_rating"

    if TEXT_COL not in df.columns:
        sys.exit(f"  ERROR: column '{TEXT_COL}' not found. "
                 f"Available columns: {list(df.columns)}")

    # keyword analysis
    print(f"\n[2/5] Analysing {len(df):,} review texts ...")
    print(f"      ({len(EXPERIENCE_KEYWORDS)} categories, "
          f"{sum(len(p) for s in EXPERIENCE_KEYWORDS.values() for p in s.values())} patterns)")

    record_list: list[dict] = []
    batch_size = 5_000

    for idx, (_, row) in enumerate(df.iterrows()):
        raw_text = row.get(TEXT_COL, "")
        text = str(raw_text) if pd.notna(raw_text) else ""
        matches = analyze_text(text)
        for cat, counts in matches.items():
            record_list.append({
                "name":     row["name"],
                "category": cat,
                "positive": counts["positive"],
                "negative": counts["negative"],
                "has_text": bool(text.strip()),
            })
        if (idx + 1) % batch_size == 0:
            print(f"      ... {idx + 1:,} rows processed")

    print(f"      Done. {len(record_list):,} keyword-match records.")

    # aggregate by attraction + category
    print(f"\n[3/5] Aggregating by attraction name ...")

    cats = list(EXPERIENCE_KEYWORDS.keys())

    if record_list:
        results_df = pd.DataFrame(record_list)
        category_df = (
            results_df
            .groupby(["name", "category"], as_index=False)
            .agg(positive=("positive", "sum"), negative=("negative", "sum"))
        )
    else:
        category_df = pd.DataFrame(columns=["name", "category", "positive", "negative"])

    summary_rows: list[dict] = []
    for name, grp in df.groupby("name", sort=False):
        texts_with_content = grp[TEXT_COL].dropna().astype(str).str.strip()
        n_with_text = int((texts_with_content != "").sum())

        avg_rating = (
            round(float(grp[RATING_COL].dropna().mean()), 2)
            if RATING_COL in grp.columns and grp[RATING_COL].notna().any()
            else None
        )

        row: dict = {
            "name":               name,
            "total_reviews":      len(grp),
            "reviews_with_text":  n_with_text,
            "avg_rating":         avg_rating,
        }

        name_cats = (
            category_df[category_df["name"] == name]
            .set_index("category")
        )
        total_pos = total_neg = 0
        for cat in cats:
            pos = int(name_cats.loc[cat, "positive"]) if cat in name_cats.index else 0
            neg = int(name_cats.loc[cat, "negative"]) if cat in name_cats.index else 0
            row[f"{cat}_pos"] = pos
            row[f"{cat}_neg"] = neg
            total_pos += pos
            total_neg += neg

        row["experience_score"] = experience_score(total_pos, total_neg)
        summary_rows.append(row)

    summary_df = (
        pd.DataFrame(summary_rows)
        .sort_values("experience_score", ascending=False, na_position="last")
        .reset_index(drop=True)
    )
    print(f"      {len(summary_df)} attractions aggregated.")

    # translations
    if args.no_translate:
        print(f"\n[4/5] Translation skipped (--no-translate flag).")
        all_patterns: list[str] = [
            p for sents in EXPERIENCE_KEYWORDS.values() for pats in sents.values() for p in pats
        ]
        translations = {p: {"English": _clean_for_translation(p)} for p in all_patterns}
    else:
        print(f"\n[4/5] Translating keywords into {len(TARGET_LANGUAGES)} languages ...")
        print(f"      Tip: use --no-translate to skip this step.\n")
        all_patterns = [
            p for sents in EXPERIENCE_KEYWORDS.values() for pats in sents.values() for p in pats
        ]
        translations = translate_keywords(all_patterns, TARGET_LANGUAGES)

    # write report
    print(f"\n[5/5] Writing Excel report ...")
    build_report(summary_df, category_df, translations, output_file)

    print("\n  -- Quick Summary --")
    top5 = summary_df.head(5)[["name", "experience_score", "total_reviews"]]
    for _, r in top5.iterrows():
        score_str = f"{r['experience_score']:.1f}" if r["experience_score"] is not None else "N/A"
        print(f"  {r['name'][:45]:<46}  score={score_str:>5}  reviews={int(r['total_reviews']):>6,}")

    print("\n  Done!")
    print("=" * 70)


if __name__ == "__main__":
    main()
